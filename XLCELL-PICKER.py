# ライブラリのインポート
import glob
import openpyxl
import csv
import pandas as pd
import os
import datetime
import hashlib
import tkinter as tk
import tkinter.simpledialog
import tkinter.filedialog
import tkinter.messagebox as messagebox
import tkinter.ttk as ttk
print("ライブラリのインポート完了")

#=================================#
# ディレクトリ選択ボタンの処理
#=================================#
#Input参照ボタンをクリックしたときの動作
def input1browse_select():
    inputfolfer_path = tk.filedialog.askdirectory(title='読込先')
    xlinput1_box.delete(0, tk.END)
    xlinput1_box.insert(tk.END, inputfolfer_path)
#Output参照ボタンをクリックしたときの動作
def outputbrowse_select():
    outputfolfer_path = tk.filedialog.askdirectory(title='出力先')
    output_box.delete(0, tk.END)
    output_box.insert(tk.END, outputfolfer_path)  
#=================================#
# 実行ボタン：本処理「action_select()」
#=================================#
def action_select():
    #------------------------------#
    #GUIからの各種値取得
    #------------------------------#
    #outputフォルダパス取得
    outputfld = output_box.get()
    #inputフォルダパス取得
    inputfld = xlinput1_box.get()
    #inputファイルパスのリスト取得
    path_str = f"{inputfld}\*.xlsx"
    paths = glob.glob(path_str)
    #チェックボックスの値取得は本処理の関数内で、出力直前にif分岐で取得する

    #------------------------------#
    #定義ファイルからの各種値取得
    #------------------------------#
    #定義ファイルから読み取り専用reade_only=Trueで値を取得
    dfnwb = openpyxl.load_workbook('定義ファイル.xlsx', read_only=True)
    dfnws = dfnwb.worksheets[0]#定義ファイルExcelの1つ目のシートを取得
    #読込対象となるExcelのシート名取得
    tgtsheetname = dfnws['B2'].value
    print(f'対象シートを「{tgtsheetname}」として取得')
    #出力時のファイル名取得
    outputfname = dfnws['B3'].value
    print(f'出力時のファイル名を「{outputfname}」として取得')

    # B列column=2から値の入っている最終行を見つける
    dfnLastrow = 11  # 開始行を一旦最終行として初期設定
    #whileループで指定した条件「セルの値がnullではない」場合、条件がTrueとなりループが継続
    while dfnws.cell(row=dfnLastrow, column=2).value is not None:
        dfnLastrow += 1#セルの値が存在する行を探すたびに1ずつ増加させる（加算代入）
    print(f'定義ファイルの最終行：{dfnLastrow - 1}行目')

    # カラム名とセル番地を取得
    dfncolumns = []#繰り返し処理の都度、定義したカラム名を順に取得するための空のリストを作成
    dfncell_addresses = []#繰り返し処理の都度、定義したセル番地を順に取得するための空のリストを作成
    for i in range(11, dfnLastrow):
        column_name = dfnws[f'B{i}'].value
        cell_address = dfnws[f'C{i}'].value
        if column_name is not None and cell_address is not None:
            dfncolumns.append(column_name)
            dfncell_addresses.append(cell_address)
        else:
            print(f'{i}行目のカラム名とセル番地がnullのためリストに追加せず')
        print(f'{i}行目まで読込完了＋カラム名「{column_name}」、セル番地「{cell_address}」を取得')

    #------------------------------#
    #対象ファイル格納先から各ファイル読込み、データフレームづくり
    #------------------------------#
    # 空のデータフレームを作成（ハッシュ値も列として入れたいところ）
    df = pd.DataFrame(columns=["ファイル名", "SHA1ハッシュ値"] + dfncolumns)

    # GUIで指定したExcel格納フォルダpaths内（*.xlsxのリスト）からに順にExcelを開き、カラムに対応する値を取得して追加していく
    for i, file_path in enumerate(paths):# file_path変数にファイルパスが1つずつ入る
        print('========================')
        print(f'{i+1}回目処理・対象ファイル:「{os.path.splitext(os.path.basename(file_path))[0]}」')
        #ハッシュ値を取得
        #withを使うことで後処理を自動で行ってくれるので、ファイルの閉じ忘れ等によるエラーを無くすことができる
        with open(file_path,'rb') as tgtfile:#rb：バイナリファイル読み取り専用モード
            r_hash = tgtfile.read()
            tgtfile_hash_sha1 = hashlib.sha1(r_hash).hexdigest()#SHA1で取得
        
        #処理回数ごとにvalue（拡張子なしのフィル名）を取得しkey(ファイル名)に紐づけ
        wb_dct = {"ファイル名": os.path.splitext(os.path.basename(file_path))[0], "SHA1ハッシュ値": tgtfile_hash_sha1}
        #paths内のExcelを読み取り専用∧計算結果の値を取得data_onlyで開く∵処理を少しでも軽くする
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        #対象シートを見つける
        sheet_exists = False #シートの有無フラグsheet_existsを初期設定としてFalseにしておくTrueにできたら処理継続できるみたいな関所
        for j, ws in enumerate(wb.worksheets):
            if ws.title == tgtsheetname:
                sheet_exists = True
                tgtsheet = wb[tgtsheetname]
                print(f'{j+1}個目のシートでヒット')
                break#対象シート見つかったらループを抜ける

        #対象シートが存在しなければその旨ポップアップ表示
        if sheet_exists == False:#「if not sheet_exists:」と同義
            print(f'指定されたシート「{tgtsheetname}」が存在しません')
            messagebox.showerror('エラー', f'ファイル「{os.path.splitext(os.path.basename(file_path))[0]}」\n内に、定義ファイルで指定したシートがありません')
        
        #カラム名のリストdfncolumnsと当該カラム名対応するセル番地リストdfncell_addressesを
        # zip関数で複数のリストの要素を同時に取得する（変数■,●in ■list,●listの対応順）
        for column, cell_address in zip(dfncolumns, dfncell_addresses):
            #定義ファイルで定義したセル番地の値を取得でvalueとして保持
            value = tgtsheet[cell_address].value
            #valueとして保持した値を対応するカラム(名)に入れる
            wb_dct[column] = value#行辞書に追加
        
        #辞書型をデータフレームに変換し、空のデータフレームに追加する用のデータフレームadd_dfとして保持
        add_df = pd.DataFrame([wb_dct])
        #空のdfに合体（縦結合axis=0(省略可)）させる
        #＋出力されるデータフレームのインデックスを連番にしたいので引数ignore_index=Trueを適用
        df = pd.concat([df, add_df], axis=0, ignore_index=True)
        wb.close()

        print(f'{i+1}回目完了')
        #プログレスバー設定
        progbar.configure(value=(i + 1)/len(paths))
        progbar.update()
    print('========================')

    #------------------------------#
    #データフレームをExcelとして出力
    #------------------------------#
    # indexを1始まりにする
    df.index = df.index + 1
    # 現在時刻取得
    dt = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    
    #チェックボックスの値（True/False）を取得しTrueならその形式で出力する
    if varxls.get() :
        # 定義ファイルで指定したディレクトリにExcelファイルとして書き出し
        df.to_excel(outputfld + f'\{outputfname}_{dt}.xlsx')    
    if varcsv.get() :
        # 定義ファイルで指定したディレクトリにcsvファイルとして書き出し
        df.to_csv(outputfld + f'\{outputfname}_{dt}.csv', encoding="shift-jis")    
    if varcsv_quotingall.get() :
        # 定義ファイルで指定したディレクトリにcsvファイルとして書き出し(ダブルコーテーション付)
        df.to_csv(outputfld + f'\{outputfname}_{dt}_quotingON.csv', encoding="shift-jis", quoting=csv.QUOTE_ALL)
    print("処理完了")
    messagebox.showinfo('メッセージ', '出力完了しました')

#=================================#
#GUIの定義
#=================================#
root = tk.Tk()
root.geometry("780x210+100+200")#画面サイズ＋左から100px上から200pxの位置にウィンドウ表示
root.title("XLCELL-PICKER Ver.1.0.0")
#Excel格納フォルダには不要なものは入れないでのメッセージラベル作成
msg1_label = tk.Label(text="Excel格納フォルダには対象のExcel以外は入れないでください。∵シート名で定義しているため")
msg1_label.place(x=10, y=10)
#読込先表示ラベルの作成
xlinput1_label = tk.Label(text="Excel格納フォルダ")
xlinput1_label.place(x=10, y=45)
#読込先表示枠の作成
xlinput1_box = tk.Entry(width=90)
xlinput1_box.place(x=110, y=45)
#読込先参照ボタンの作成
button1 = tk.Button(text="参照",width=10,command = input1browse_select)
button1.place(x=680, y=42)
#出力先表示枠の作成
output_box = tk.Entry(width=90)
output_box.place(x=110, y=75)
#出力先表示ラベルの作成
output_label = tk.Label(text="出力先フォルダ")
output_label.place(x=10, y=75)
#出力先参照ボタンの作成
button2 = tk.Button(text="参照",width=10,command=outputbrowse_select)
button2.place(x=680, y=72)

#チェックボックスの設定
#出力ファイル形式選択の表示ラベルの作成
outputtype_label = tk.Label(text="出力ファイル形式")
outputtype_label.place(x=10, y=103)
#チェックボックス1（Excel形式）
varxls = tk.BooleanVar()
varxls.set( True ) #初期値をTrueに設定
che1 = tk.Checkbutton( text = '.xlsxファイル', variable = varxls )
che1.place(x=108, y=100)
#チェックボックス2（csv形式）
varcsv = tk.BooleanVar()
varcsv.set( True ) #初期値をTrueに設定
che2 = tk.Checkbutton( text = '.csvファイル', variable = varcsv )
che2.place(x=200, y=100)
#チェックボックス3（csv形式かつ項目名含め全値にダブルコーテーション付）
varcsv_quotingall = tk.BooleanVar()
varcsv_quotingall.set( True ) #初期値をTrueに設定
che3 = tk.Checkbutton( text = '.csvファイル(ダブルクォーテーション付き)', variable = varcsv_quotingall )
che3.place(x=300, y=100)


#定義ファイルへの記入を促すラベル作成
msg2_label = tk.Label(text="当該toolと同階層にある「定義ファイル.xlsx」で諸々入力・上書き保存したうえで実行してください")
msg2_label.place(x=10, y=130)
#実行ボタンの作成
Button=tk.Button(text="実行",width=10,command=action_select)
Button.place(x=680, y=130)
# プログレスバー配置
progbar = ttk.Progressbar(root, length=740, mode="determinate", maximum=1)
progbar.place(x=20, y=170)

root.mainloop()#ウィンドウ表示
#------------------------------#
